# backend/seed_from_excel.py
from __future__ import annotations
from openpyxl import load_workbook
from pathlib import Path
from typing import Optional, Tuple
from db import Base, engine, SessionLocal
from models import User

# ====== คีย์ที่รองรับ (ไม่สนช่องว่าง/จุด/ขีดล่าง/ตัวพิมพ์) ======
def norm(s: str) -> str:
    return (s or "").strip().lower().replace(".", "").replace(" ", "").replace("_", "")

EMP_KEYS    = {"employeeid", "empid", "id", "employeeno", "employeenumber"}
NAME_KEYS   = {"name", "displayname", "fullname"}
FIRST_KEYS  = {"firstname", "first", "given", "givenname"}
LAST_KEYS   = {"lastname", "last", "surname", "familyname"}
EMAIL_KEYS  = {"email", "mail"}
DEPT_KEYS   = {"department", "dept", "costcenter", "org", "organization"}
AVATAR_KEYS = {"avatarurl", "avatar", "photo", "photourl", "picture"}

def detect_header_row(ws, max_seek_rows: int = 10) -> int:
    """ไล่หาแถวหัวคอลัมน์ภายใน 10 แถวแรก (เจอคำว่า employee ที่หัวคอลัมน์ใดก็ได้)"""
    for r in range(1, max_seek_rows + 1):
        row_vals = [str(c.value or "") for c in ws[r]]
        keys = [norm(v) for v in row_vals]
        if any("employee" in k for k in keys):
            return r
    return 1

def find_idx(keys_norm: list[str], candidates: set[str], default: int = -1) -> int:
    for i, k in enumerate(keys_norm):
        if k in candidates:
            return i
    return default

def to_digits(raw) -> str:
    """แปลงค่า cell เป็นเลขล้วน (ตัด EN/ตัวอักษร/ทศนิยม)"""
    if raw is None:
        return ""
    s = str(raw).strip()
    s = s.upper().removeprefix("EN")
    # เผื่อกรณีมาเป็น float 102562.0 → 102562
    if s.endswith(".0"):
        s = s[:-2]
    return "".join(ch for ch in s if ch.isdigit())

def seed_from_excel(xlsx_path: str = "datauser.xlsx", sheet_name: Optional[str] = None):
    p = Path(xlsx_path)
    if not p.exists():
        print("Excel not found:", p.resolve()); return

    Base.metadata.create_all(bind=engine)

    wb = load_workbook(p, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    header_row = detect_header_row(ws)
    headers = [str(c.value or "").strip() for c in ws[header_row]]
    keys_norm = [norm(h) for h in headers]

    # ตำแหน่งคอลัมน์
    i_emp   = find_idx(keys_norm, EMP_KEYS, default=0)  # ถ้าไม่เจอ ใช้คอลัมน์แรกเป็นรหัส
    i_name  = find_idx(keys_norm, NAME_KEYS, -1)
    i_first = find_idx(keys_norm, FIRST_KEYS, -1)
    i_last  = find_idx(keys_norm, LAST_KEYS, -1)
    i_email = find_idx(keys_norm, EMAIL_KEYS, -1)
    i_dept  = find_idx(keys_norm, DEPT_KEYS, -1)
    i_avt   = find_idx(keys_norm, AVATAR_KEYS, -1)

    print(f"[seed] sheet='{ws.title}' header_row={header_row}")
    print(f"[seed] columns: {headers}")

    added = skipped = 0
    examples: list[Tuple[str, str]] = []

    with SessionLocal() as db:
        for row in ws.iter_rows(min_row=header_row + 1):
            # ดึงค่าในแต่ละคอลัมน์อย่างปลอดภัย
            get = lambda idx: (str(row[idx].value).strip() if 0 <= idx < len(row) and row[idx].value is not None else "")

            emp = to_digits(get(i_emp))
            if not (6 <= len(emp) <= 7):
                skipped += 1
                continue

            # name: ใช้ name → หรือ first + last
            name = get(i_name)
            if not name:
                first = get(i_first)
                last  = get(i_last)
                name = (first + " " + last).strip() or "Unknown"

            email = get(i_email) or None
            dept  = get(i_dept) or None
            avatar= get(i_avt) or None

            # ซ้ำข้าม
            if db.query(User).filter(User.employee_id == emp).first():
                continue

            u = User(employee_id=emp, name=name, email=email, department=dept, avatar_url=avatar)
            db.add(u); added += 1
            if len(examples) < 5:
                examples.append((emp, name))

        db.commit()

    print(f"[seed] added={added}, skipped={skipped}")
    if examples:
        print("[seed] examples:", examples)

if __name__ == "__main__":
    # เปลี่ยนชื่อไฟล์/ชีทได้ตามจริง เช่น seed_from_excel('employees.xlsx', 'Active')
    seed_from_excel("datauser.xlsx", None)
